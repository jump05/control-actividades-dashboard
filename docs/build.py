# -*- coding: utf-8 -*-
"""
Lee ../base de datos/*.xlsx y genera data.js para el dashboard simplificado
de Control de Actividades (Mtto / TI).

Logica (simplificada a pedido):
  - Por cada (colaborador, dia) se toma solo la PRIMERA sede reportada ese dia
    en Base_Mtto/Base_TI (la primera fila que aparece para ese dia).
  - Se compara contra la PRIMERA marcacion biometrica real del dia (Entrada 1 /
    dispositivo de entrada 1), no todas las marcaciones del dia.
  - Si el dispositivo de la primera marcacion no corresponde a la sede
    reportada (o no hay marcacion ese dia), el dia cuenta como "incidencia".
  - Dias de ausencia (Vacaciones/Feriado/etc.), fuera del rango de marcaciones,
    sin DNI identificado, o en sedes sin dispositivo biometrico, no se cuentan
    (no se pueden verificar).
"""
import json
import re
import unicodedata
from datetime import datetime, date
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
DB = HERE.parent / "base de datos"
MAESTRO = DB / "Maestro Control Actividades.xlsx"
MARCACIONES = DB / "MARCACIONES INFRA-MMTTO-TI.xlsx"
CONTROL = DB / "Control de actividades - Mtto y TI.xlsx"

# palabra(s) clave que identifican de forma unica a cada sede dentro del
# nombre del dispositivo biometrico (evita depender de "Biometrico -" que
# no aparece en todas las variantes, p.ej. "Dasso 5to Piso I")
KEYWORDS = {
    "La Victoria": ["VICTORIA"],
    "Colina": ["COLINA"],
    "Faucett": ["FAUCETT"],
    "Surquillo": ["SURQUILLO"],
    "Surco": ["SURCO"],
    "Independencia": ["INDEPENDENCIA"],
    "San Juan de Lurigancho": ["SJL"],
    "Dasso": ["DASSO"],
    "Santa Anita": ["SANTA", "ANITA"],
    "San Borja (Eureka)": ["SAN", "BORJA"],
    "Derby": ["DERBY"],
    "Ate": ["ATE"],
    "Treneman": ["TRENEMAN"],
    "El Polo (Eureka)": ["POLO"],
    "Chorrillos": ["CHORRILLOS"],
    "San Miguel": ["MIGUEL"],
}


def norm(s):
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.upper()


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def sede_for_device(device_norm):
    for sede, kws in KEYWORDS.items():
        if all(k in device_norm for k in kws):
            return sede
    return None


def device_matches_sede(sede_friendly, device_norm):
    kws = KEYWORDS.get(sede_friendly)
    if not kws:
        return False
    return all(k in device_norm for k in kws)


# ------------------------------------------------------------------
# 1) MAESTRO
# ------------------------------------------------------------------
wb_m = openpyxl.load_workbook(MAESTRO, data_only=True)

trabajadores = {}
for row in wb_m["Trabajadores"].iter_rows(min_row=2, values_only=True):
    dni, nombre, cargo, area = row[0], row[1], row[2], row[3]
    if not dni:
        continue
    trabajadores[str(dni).strip()] = {"nombre": nombre, "cargo": cargo, "area": area}

alias_to_dni = {}
for row in wb_m["Alias_Colaborador"].iter_rows(min_row=2, values_only=True):
    aliasv, dni = row[0], row[1]
    if not aliasv:
        continue
    alias_to_dni[norm(aliasv)] = str(dni).strip() if dni else None

sedes_info = {}  # abreviado normalizado -> {sede_friendly, tiene_bio}
for row in wb_m["Sedes"].iter_rows(min_row=2, values_only=True):
    abrev, sede, disp, tiene, nota = row[0], row[1], row[2], row[3], row[4]
    if not abrev:
        continue
    sedes_info[norm(abrev)] = {
        "sede": sede,
        "tiene_bio": str(tiene).strip().lower().startswith("s"),
    }

ausencias = set()
for row in wb_m["Ausencias_NoSede"].iter_rows(min_row=2, values_only=True):
    if row[0]:
        ausencias.add(norm(row[0]))

# ------------------------------------------------------------------
# 2) MARCACIONES -> primera marcacion del dia por (dni, fecha)
# ------------------------------------------------------------------
wb_k = openpyxl.load_workbook(MARCACIONES, data_only=True)
ws_k = wb_k["CONTROL REGISTRO DE ASISTENCIA"]

primera_marcacion = {}  # (dni, fecha) -> {marco: bool, device: str norm}
marc_fechas = []

for row in ws_k.iter_rows(min_row=2, values_only=True):
    fecha = to_date(row[0])
    if fecha is None:
        continue
    dni = str(row[1]).strip() if row[1] else None
    if not dni:
        continue
    marc_fechas.append(fecha)
    horario1 = str(row[4]).strip() if row[4] else "----"
    entrada1 = str(row[5]).strip() if row[5] else "---"
    device1 = str(row[6]).strip() if row[6] else ""
    marco = horario1 != "----" and entrada1 not in ("---", "")
    primera_marcacion[(dni, fecha)] = {
        "marco": marco,
        "device": norm(device1) if marco else "",
    }

MARC_MIN = min(marc_fechas) if marc_fechas else None
MARC_MAX = max(marc_fechas) if marc_fechas else None

# ------------------------------------------------------------------
# 3) BASE_MTTO / BASE_TI -> primera sede reportada por (dni, fecha)
# ------------------------------------------------------------------
wb_c = openpyxl.load_workbook(CONTROL, data_only=True)

primer_reporte = {}  # (dni_or_alias, fecha) -> {dni, alias, area_hint, sede_raw}


def procesar(sheetname, area_hint):
    ws = wb_c[sheetname]
    for row in ws.iter_rows(min_row=2, values_only=True):
        fecha, mes, colab, unidad, sede_raw = row[1], row[2], row[3], row[4], row[5]
        fecha = to_date(fecha)
        if not colab or fecha is None:
            continue
        alias_key = norm(colab)
        dni = alias_to_dni.get(alias_key)
        key = (dni or alias_key, fecha)
        if key in primer_reporte:
            continue  # ya se tomo la primera fila de ese dia
        primer_reporte[key] = {
            "dni": dni, "alias": colab, "area_hint": area_hint,
            "sede_raw": sede_raw, "unidad": unidad,
        }


procesar("Base_Mtto", "Mantto")
procesar("Base_TI", "TI")

# ------------------------------------------------------------------
# 4) evaluar cada dia
# ------------------------------------------------------------------
por_persona = {}  # dni -> {area, nombre, dias:0, incidencia:0}
marcas_persona = {}  # dni -> Counter(sede_friendly)

for (dni_or_alias, fecha), rep in primer_reporte.items():
    dni = rep["dni"]
    if not dni or dni not in trabajadores:
        continue  # colaborador no identificado (ej. Ricardo braul) -> no se cuenta

    sede_raw = rep["sede_raw"]
    sede_key = norm(sede_raw)
    es_ausencia = sede_key in ausencias or norm(rep["unidad"]) in ausencias
    if es_ausencia:
        continue
    if MARC_MIN and (fecha < MARC_MIN or fecha > MARC_MAX):
        continue

    info = sedes_info.get(sede_key)
    if not info or not info["tiene_bio"]:
        continue  # sede sin dispositivo biometrico -> no verificable

    mk = primera_marcacion.get((dni, fecha))
    confirmado = bool(mk and mk["marco"] and device_matches_sede(info["sede"], mk["device"]))

    trab = trabajadores[dni]
    p = por_persona.setdefault(dni, {"area": trab["area"], "nombre": trab["nombre"], "dias": 0, "incidencia": 0})
    p["dias"] += 1
    if not confirmado:
        p["incidencia"] += 1

# distribucion de sedes (donde marca cada personal), en base a TODAS sus
# primeras marcaciones reales dentro del rango de Marcaciones (no solo los
# dias con actividad reportada) para reflejar su patron real de asistencia
dnis_tabla = set(por_persona.keys())
for (dni, fecha), mk in primera_marcacion.items():
    if dni not in dnis_tabla or not mk["marco"]:
        continue
    sede_friendly = sede_for_device(mk["device"]) or (mk["device"].title() if mk["device"] else "Otro")
    marcas_persona.setdefault(dni, {})
    marcas_persona[dni][sede_friendly] = marcas_persona[dni].get(sede_friendly, 0) + 1

# ------------------------------------------------------------------
# 5) salida
# ------------------------------------------------------------------
resumen = []
for dni, p in por_persona.items():
    pct = round(p["incidencia"] / p["dias"] * 100) if p["dias"] else 0
    resumen.append({
        "area": p["area"], "dni": dni, "personal": p["nombre"],
        "dias": p["dias"], "incidencia": p["incidencia"], "porcentaje": pct,
    })
resumen.sort(key=lambda r: (-r["porcentaje"], r["personal"]))

sede_chart = []
for dni, counts in marcas_persona.items():
    total = sum(counts.values())
    top = sorted(counts.items(), key=lambda x: -x[1])
    sede_chart.append({
        "dni": dni, "personal": trabajadores[dni]["nombre"], "area": trabajadores[dni]["area"],
        "total": total,
        "sedes": [{"sede": s, "dias": n, "pct": round(n/total*100)} for s, n in top],
    })
sede_chart.sort(key=lambda r: r["personal"])

out = {
    "generatedAt": datetime.now().isoformat(timespec="seconds"),
    "rangoMarcaciones": {
        "min": MARC_MIN.isoformat() if MARC_MIN else None,
        "max": MARC_MAX.isoformat() if MARC_MAX else None,
    },
    "resumen": resumen,
    "sedeChart": sede_chart,
}

version = datetime.now().strftime("%Y%m%d%H%M%S")
js = "const DASH_DATA = " + json.dumps(out, ensure_ascii=False, indent=None) + ";\n"
(HERE / "data.js").write_text(js, encoding="utf-8")

idx_path = HERE / "index.html"
if idx_path.exists():
    html = idx_path.read_text(encoding="utf-8")
    html = re.sub(r'(data\.js|app\.js)\?v=\d+', lambda m: f"{m.group(1)}?v={version}", html)
    idx_path.write_text(html, encoding="utf-8")

print(f"OK -> {len(resumen)} personas, rango marcaciones {MARC_MIN} a {MARC_MAX}")
for r in resumen:
    print(f"   {r['area']:7} {r['dni']:10} {r['personal']:38} dias={r['dias']:3} incidencia={r['incidencia']:3} ({r['porcentaje']}%)")
print(f"data.js escrito, version {version}")
